#
# Result Class
#
import openpyxl


class Result:
    def __init__(self):
        self.file_name = 'result/result.xlsx'
        self.wb = openpyxl.load_workbook(self.file_name)
        self.sheet = self.wb['result']
        print(type(self.sheet))

    def save_sheet(self):
        self.wb.save(self.file_name)

    def get_row(self, id):
        for row in self.sheet.iter_rows(min_row=2):
            if row[0].value == id:
                return row
            else:
                continue

    def add_data(self, id, participant_name, fixation_ave_x, fixation_ave_y):
        for row in self.sheet.iter_rows(min_row=2):
            if row[0].value == id:
                col_numbers = [1, 4, 7, 10, 13]
                for col_number in col_numbers:
                    print(row[col_number].value)
                    print(participant_name)
                    print(fixation_ave_x)
                    print(fixation_ave_y)
                    if row[col_number].value == None:
                        row[col_number].value = participant_name
                        row[col_number+1].value = fixation_ave_x
                        row[col_number+2].value = fixation_ave_y
                        break

            else:
                continue
